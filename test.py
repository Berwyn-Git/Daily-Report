import openpyxl
import re
from datetime import datetime, timedelta

names_positions = {'Lu Yongjin': 'Spare part keeper',
                   'Xu Zhiliang': 'Shift leader',
                   'Yuan Jiqun': 'Straightening & Cutting',
                   'Zhu Zhengkai': 'Bending',
                   'Gao Qi': 'Bending',
                   'Zhao Shuangyong': 'Bending',
                   'Chen Ming': 'Filter Deputy Supervisor',
                   'Yu Wenbin': 'Filter',
                   'Hu Ming': 'Filter',
                   'Wu Wangchao': 'Filter',
                   'Ren Chaofeng': 'Electrican Supervisor',
                   'Shen Weiguo': 'Electrican',
                   'Tang Jiabang': 'Electrican',
                   'Hu Maozhen': 'Electrican',
                   'Fang Guozhao': 'Heat Treatment Group Leader',
                   'Xu Ming': 'Heat Treatment Group Leader',
                   'Yao Liangjun': 'Heat Treatment Group Leader',
                   'Li Yong': 'Heat Treatment Operator',
                   'Zhang Chunming': 'Heat Treatment Operator',
                   'Liu Yunping': 'Heat Treatment Operator',
                   'Liu Ren': 'Pickling',
                   'Xia Feng': 'Boiler',
                   'Zhai Ming': 'Boiler',
                   'Huang Changyin': 'Boiler',
                   'Meng Xiangying': 'Boiler',
                   'Jiang Chuanyao': 'Facility',
                   'Huang Saisai': 'Facility',
                   'Zhang Kechun': 'Facility',
                   'Zhang Zihao': 'Facility',
                   'Ding Derun': 'Calibrating',
                   'Liang Tao': 'Dispatch Assistant',
                   'Fang Qinlong': 'Dispatch',
                   'Zhai Xiaorui': 'Dispatch',
                   'Zhang Zhenping': 'Dispatch',
                   'Xie Tingkun': 'Dispatch',
                   'Shi Wenbin': 'Dispatch',
                   'Zhang Doqing': 'Dispatch',
                   'Yan Yuanyang': 'Dispatch',
                   'Dong Xianglong': 'Dispatch',
                   'Dai Yanchao': 'Dispatch',
                   'He Mengjie': 'Dispatch',
                   'Ai Hongxia': 'Dispatch',
                   'He Ligang': 'Dispatch',
                   'Zhang Xuliang': 'Dispatch',
                   'Liang Yapeng': 'Dispatch',
                   'Cheng Lianlian': 'Dispatch'}

def extract_data_weekend():
    workbook = openpyxl.load_workbook('ShiftplanFebruary2.xlsx')
    sheet = workbook['Sheet1']
    new_workbook = openpyxl.load_workbook('shiftplan_extracted.xlsx')
    new_sheet = new_workbook.active

    for area in sheet.iter_rows(min_row=20, max_col=10, max_row=sheet.max_row):
        for cell in area[2:3]:
            cell_value = cell.value
            cellpattern = re.compile(r'.*[EMN].*')
            shift_table = {
                'E': '07:00 - 15:30',
                'M': '15:00 - 00:00',
                'N': '00:00 - 07:00'
            }
            if cellpattern.match(str(cell_value)):
                employee_name = sheet.cell(row=cell.row, column=1).value
                english_name = ' '.join(re.findall(r'[A-Z][a-z]+', employee_name))
                if english_name in names_positions:
                    position = names_positions[english_name]
                else:
                    position = ''
                employee_number = sheet.cell(row=cell.row, column=2).value
                OT_date = sheet.cell(row=19, column=cell.column).value
                ot_type = '双休Weekend'
                shift_target = re.findall(r'[EMN]', str(cell_value))
                if shift_target is not None and shift_target:
                    shift = shift_target[0]
                    if shift in shift_table:
                        start_time = shift_table[shift].split(' - ')[0]
                        end_time = shift_table[shift].split(' - ')[1]
                    else:
                        start_time = ""
                        end_time = ""
                else:
                    shift = ""
                    start_time = ""
                    end_time = ""
                    start_time = shift_table[shift_target].split(' - ')[0]
                    start_time_revised = datetime.strptime(start_time, '%H:%M')
                    end_time = shift_table[shift_target].split(' - ')[1]
                    end_time_revised = datetime.strptime(end_time, '%H:%M')

                    ot_hours = (end_time_revised - start_time_revised).seconds/3600

                new_sheet.append([english_name, employee_number, position, OT_date, shift, start_time, end_time, ot_type, ot_hours, '', '', '', '', ''])
    new_workbook.save('shiftplan_extracted.xlsx')

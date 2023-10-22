import random
import openpyxl
from openpyxl import load_workbook
import re
import os
from datetime import datetime, timedelta
from time import sleep
from tqdm import tqdm
import tkinter as tk
from tkinter import *
from tkinterdnd2 import *


def get_file_path():
    win = TkinterDnD.Tk()
    win.title('Drag your file here!请把文件拖拽到这里！')
    win.config(bg='gold')

    def path_listbox(event):
        file_path = event.data.strip()
        listbox.insert("end", file_path)
        win.quit()

    frame = Frame(win)
    frame.pack()

    listbox = Listbox(
        frame,
        width=75,
        height=15,
        selectmode=SINGLE
    )
    listbox.pack(fill=X, side=LEFT)
    listbox.drop_target_register(DND_FILES)
    listbox.dnd_bind('<<Drop>>', path_listbox)

    scrolbar = Scrollbar(frame, orient=VERTICAL)
    scrolbar.pack(side=RIGHT, fill=Y)

    listbox.configure(yscrollcommand=scrolbar.set)
    scrolbar.config(command=listbox.yview)

    win.mainloop()
    file_path = listbox.get("0")
    win.destroy()
    return file_path


file_path = get_file_path()
print(file_path)


def prepare_data():
    # Create a dictionary to store the names and positions
    global names_positions, workbook, sheet, new_workbook_shiftplan, new_sheet_shiftplan, new_workbook_standby, new_sheet_standby, vd
    vd = datetime.now()
    names_positions = {'Lu Yongjin': 'Spare part keeper',
                   'Xu Zhiliang': 'Shift leader',
                   'Yuan Jiqun': 'Straightening & Cutting',
                   'Zhu Zhengkai': 'Bending',
                   'Gao Qi': 'Bending',
                   'Zhao Shuangyong': 'Bending',
                   'Chen Ming': 'Filter Deputy Supervisor',
                   'Yu Wenbin': 'Filter',
                   'Hu Ming': 'Filter',
                   'Wu Wangchao': 'Filter',
                   'Ren Chaofeng': 'Electrican Supervisor',
                   'Shen Weiguo': 'Electrican',
                   'Tang Bangjia': 'Electrican',
                   'Hu Maodian': 'Electrican',
                   'Fang Guozhao': 'Heat Treatment Group Leader',
                   'Xu Ming': 'Heat Treatment Group Leader',
                   'Yao Liangjun': 'Heat Treatment Group Leader',
                   'Li Yong': 'Heat Treatment Operator',
                   'Zhang Chunming': 'Heat Treatment Operator',
                   'Liu Yunping': 'Heat Treatment Operator',
                   'Liu Ren': 'Pickling',
                   'Xia Feng': 'Boiler',
                   'Zhai Ming': 'Boiler',
                   'Huang Changyin': 'Boiler',
                   'Meng Xiangying': 'Boiler',
                   'Jiang Chuanyao': 'Facility',
                   'Huang Saisai': 'Facility',
                   'Zhang Kechun': 'Facility',
                   'Zhang Zihao': 'Facility',
                   'Ding Derun': 'Calibrating',
                   'Liang Tao': 'Dispatch Assistant',
                   'Fang Qinlong': 'Dispatch',
                   'Zhai Xiaorui': 'Dispatch',
                   'Zhang Zhenping': 'Dispatch',
                   'Xie Tingkun': 'Dispatch',
                   'Shi Wenbin': 'Dispatch',
                   'Zhang Doqing': 'Dispatch',
                   'Yan Yuanyang': 'Dispatch',
                   'Dong Xianglong': 'Dispatch',
                   'Dai Yanchao': 'Dispatch',
                   'He Mengjie': 'Dispatch',
                   'Ai Hongxia': 'Dispatch',
                   'He Ligang': 'Dispatch',
                   'Zhang Xuliang': 'Dispatch',
                   'Liang Yapeng': 'Dispatch',
                   'Cheng Lianlian': 'Dispatch'}
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the worksheet
    sheet = workbook['Sheet1']

    # Create a new workbook to store the extracted data
    new_workbook_shiftplan = openpyxl.Workbook()
    new_sheet_shiftplan = new_workbook_shiftplan.active

    new_workbook_standby = openpyxl.Workbook()
    new_sheet_standby = new_workbook_standby.active

    # Add the headers to the new worksheet
    new_sheet_shiftplan['A1'] = '姓名 Name'
    new_sheet_shiftplan['B1'] = '工号 Employee No.'
    new_sheet_shiftplan['C1'] = '岗位 Position'
    new_sheet_shiftplan['D1'] = '预计加班日期 Plan OT Date'
    new_sheet_shiftplan['E1'] = '班次 Shift'
    new_sheet_shiftplan['F1'] = '预计加班时间从 Plan OT from'
    new_sheet_shiftplan['G1'] = '到 To'
    new_sheet_shiftplan['H1'] = '加班类型 OT Type'
    new_sheet_shiftplan['I1'] = '小时数 Hrs'
    new_sheet_shiftplan['J1'] = '用餐时间 Meal Time'
    new_sheet_shiftplan['K1'] = '加班时数 OT Hours'
    new_sheet_shiftplan['L1'] = '事由（请详细描述） Working Reason (Please clarify in detail)'
    new_sheet_shiftplan['M1'] = '补偿方式 Way of OT settlement'
    new_sheet_shiftplan['N1'] = '申请人签名 Applicant Signature'

    new_sheet_standby['A1'] = '姓名 Name'
    new_sheet_standby['B1'] = '工号 Number'
    new_sheet_standby['C1'] = '岗位 Position'
    new_sheet_standby['D1'] = '预计待命日期 Plan Standby Date'
    new_sheet_standby['E1'] = '预计待命起讫时间 Plan Standby Period/Hours'
    new_sheet_standby['F1'] = '事由 Working Reason'
    new_sheet_standby['G1'] = '实际待命起讫时间 Actual Standby Period/Hours'
    new_sheet_standby['H1'] = '待命时数 Standby Hours'


def change_columns():
    # Create a tlinter window
    trigger_window = tk.Tk()

    # Create a label and entry for input number
    label = tk.Label(trigger_window, height=35, text="Enter the total columns you need(请输入本次导入文件的总列数)", width= 70)
    label.pack()
    entry = tk.Entry(trigger_window, width=35)
    entry.pack()

    def change_trigger():
        # Create a button to trigger the data changing function
        global columns_number
        columns_number = int(entry.get())
        extract_data()

    button = tk.Button(trigger_window, text="Confirm Input(确认并执行)", command=change_trigger)
    button.pack()
    trigger_window.mainloop()


def extract_data():
    # Iterate over the rows in the sheet
    for area in sheet.iter_rows(min_row=20, max_col=columns_number, max_row=sheet.max_row):
        # Iterate over the cells in the row
        for cell in area[2:9]:
            cell_value = cell.value
            # Check if the cell value matches the desired format
            cellpattern = re.compile(r'.*\d{1,2}:\d{2}[ -]*\d{1,2}:\d{2}.*')
            # if match:
            if cell_value is not None and cellpattern.match(str(cell_value)):
                # Get the employee name from column A
                employee_name = sheet.cell(row=cell.row, column=1).value
                # Extract the English part of the name
                english_name = ' '.join(re.findall(r'[A-Z][a-z]+', employee_name))

                # Get the employee number from column B
                employee_number = sheet.cell(row=cell.row, column=2).value

                # Check if the employee name is in the names_and_positions dictionary
                if english_name in names_positions:
                    # Get the position for the employee
                    position = names_positions[english_name]
                else:
                    position = ""

                # Get the corresponding date for this cell from row 19
                OT_date = sheet.cell(row=1, column=cell.column).value
                OT_date_final = datetime.strptime(OT_date.strftime('%Y-%m-%d'), '%Y-%m-%d')

                times = cell.value.split(' - ')

                if OT_date_final.weekday() >= 5:
                    ot_type = '双休Weekend'
                    if len(times) == 2:
                        start_time_mixed, end_time_mixed = times
                        start_time = ' '.join(re.findall(r'\d{1,2}:\d{2}', start_time_mixed))
                        end_time = ' '.join(re.findall(r'\d{1,2}:\d{2}', end_time_mixed))
                        # weekend all the time should be calculated as overtime
                        ot_hours = (datetime.strptime(end_time, '%H:%M') - datetime.strptime(start_time,
                                                                                             '%H:%M')).seconds / 3600
                        # Determine the shift based on the start time
                        if start_time in ['07:00', '08:30']:
                            shift = 'E'
                        elif start_time == '15:30':
                            shift = 'M'
                        elif start_time in ['19:00', '00:00']:
                            shift = 'N'
                        else:
                            shift = ''

                else:
                    ot_type = '平时Weekday'
                    if len(times) == 2:
                        start_time_mixed, end_time_mixed = times
                        start_time = ' '.join(re.findall(r'\d{1,2}:\d{2}', start_time_mixed))
                        end_time = ' '.join(re.findall(r'\d{1,2}:\d{2}', end_time_mixed))

                        # Determine the shift based on the start time
                        if start_time in ['07:00', '08:30']:
                            shift = 'E'
                        elif start_time == '15:30':
                            shift = 'M'
                        elif start_time in ['19:00', '21:00', '00:00']:
                            shift = 'N'
                        else:
                            shift = ''

                        if shift in ['E', 'M']:
                            supposed_end_time_datetime = (
                                datetime.strptime(start_time, '%H:%M') + timedelta(hours=8.5)).strftime('%H:%M')
                            if supposed_end_time_datetime != end_time:
                                start_time = supposed_end_time_datetime
                                # Get the ot hours
                                ot_hours = (datetime.strptime(end_time, '%H:%M') - datetime.strptime(start_time, '%H:%M')).seconds / 3600
                            else:
                                ot_hours = 8.5
                        if shift == 'N':
                            supposed_start_time_datetime = (
                                datetime.strptime(end_time, '%H:%M') - timedelta(hours=7.0)).strftime('%H:%M')
                            if supposed_start_time_datetime != start_time:
                                end_time = supposed_start_time_datetime
                                ot_hours = (datetime.strptime(end_time, '%H:%M') - datetime.strptime(start_time, '%H:%M')).seconds / 3600
                            else:
                                ot_hours = 7.0
                    else:
                        cell_row = cell.row
                        print(cell_row)

                # Add the data to the new worksheet
                new_sheet_shiftplan.append(
                    [english_name, employee_number, position, OT_date_final, shift, start_time, end_time, ot_type,
                     ot_hours, '', '', '', '', ''])


def extract_standby_data():
    global vs
    vs = "2023-7-30"
    for area in sheet.iter_rows(min_row=20, max_col=9, max_row=sheet.max_row):
        for cell in area[2:9]:
            cell_value = cell.value
            cellpattern = re.compile(r'.*\s(\d{1,2}\s*h).*')
            if cell_value is not None and cellpattern.match(str(cell_value)):
                employee_name = sheet.cell(row=cell.row, column=1).value
                english_name = ' '.join(re.findall(r'[A-Z][a-z]+', employee_name))
                employee_number = sheet.cell(row=cell.row, column=2).value
                if english_name in names_positions:
                    # Get the position for the employee
                    position = names_positions[english_name]
                else:
                    position = ""
                standby_date = sheet.cell(row=1, column=cell.column).value
                standby_hours = re.findall(r'\d+h', cell_value)[0]

                new_sheet_standby.append(
                    [employee_name, employee_number, position, standby_date, standby_hours]
                )


def sort_data():
    # Save the new workbook
    new_workbook_shiftplan.save('sample_extracted.xlsx')
    new_workbook_standby.save('standby_extracted.xlsx')


print("在本程序开始之前，首先请检查以下步骤是否完成：")
print("1. 将本周所需操作的时间段从原先的excel文件中提取出来，即从周六周日开始到本周五结束")
print("2. 所有的员工姓名的英文部分请按照‘Zhang Bowen’这样的格式：姓与名的第一个字母大写")
print("3. 将所有的加班内容里所有的班次全部转化为数字内容，例如‘19:00 - 07:00’这种样式")
print("4. 所有含有例如‘19:00 - 07:00’的单元格中不得出现除时间分割符以外的其他‘-’分割符")
status = input("是否完成以上步骤：(如果是请输入Y，如果不是请输入N，并完成以上步骤)")

# if status == 'Y':
#     prepare_data()
#     extract_data()
#     extract_standby_data()
#     sort_data()
#
#
#     def my_function():
#         for i in range(10):
#             sleep(random.uniform(0.0001, 0.01))
#
#
#     for i in tqdm(range(100), desc="Processing", ascii=False, ncols=75):
#         my_function()
#
#     print("\n所有工作已经完成！请检查输出的结果文件！")
# else:
#     print("\n请准备好再运行本程序")

error_codes = {
    "'NoneType' object has no attribute 'strftime'":"\n日期的读取出现了问题，检查表格文件中日期行是否是第1行",
    "openpyxl does not support .xlsx} file format, please check you can open it with Excel first. Supported formats are: .xlsx,.xlsm,.xltx,.xltm":"\n导入文件的文件名包含空格，请将文件名的空格全部删除！"
}

try:
    if status == 'Y' and vd <= vs:
        prepare_data()
        change_columns()
        extract_data()
        extract_standby_data()
        sort_data()


        def my_function():
            for i in range(10):
                sleep(random.uniform(0.0001, 0.01))


        for i in tqdm(range(100), desc="Processing", ascii=False, ncols=75):
            my_function()
        print("\n所有工作已经完成！请检查输出的结果文件！")
    else:
        print("\n请准备好再运行本程序")
except Exception as e:
    error_message = str(e)
    print(error_message)
    if error_message in error_codes:
        print(error_codes[error_message])
    else:
        print("\n未记录的错误出现，请查看源代码")

folder_path = r"C:\Users\bowen.zhang\PycharmProjects\pythonProject"
os.system(f"start explorer {folder_path}")